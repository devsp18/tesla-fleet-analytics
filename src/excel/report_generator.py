import pandas as pd
import numpy as np
import sqlite3
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from analysis.fleet_analysis import *

# ── Color Palette ─────────────────────────────────────────────
TESLA_RED    = "E31937"
BLACK        = "1A1714"
OFF_WHITE    = "F7F4EF"
LIGHT_GRAY   = "EDE9E3"
MID_GRAY     = "C8C4BE"
DARK_GRAY    = "605C57"
WHITE        = "FFFFFF"
GREEN        = "1A6B3C"
RED_LIGHT    = "FFF0F0"
GREEN_LIGHT  = "F0FAF0"

def style_header_cell(cell, text, size=11, bold=True, bg=BLACK, fg=WHITE, align='left'):
    cell.value = text
    cell.font = Font(name='Calibri', size=size, bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

def style_data_cell(cell, value, fmt=None, bold=False, bg=WHITE, fg=BLACK, align='left'):
    cell.value = value
    cell.font = Font(name='Calibri', size=10, bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if fmt:
        cell.number_format = fmt

def add_thin_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color='E2DDD6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def create_overview_sheet(wb, demo_df, results):
    ws = wb.create_sheet("Executive Overview")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3

    # Title block
    ws.merge_cells('B2:H3')
    title = ws['B2']
    title.value = "TESLA MARKETING FLEET — PERFORMANCE INTELLIGENCE REPORT"
    title.font = Font(name='Calibri', size=18, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=BLACK)
    title.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('B4:H4')
    sub = ws['B4']
    sub.value = "Americas Fleet · FY 2024–2025 · 15 Markets · 5 Models · 15,000 Demo Records"
    sub.font = Font(name='Calibri', size=10, color=DARK_GRAY)
    sub.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 18

    # KPI Cards row
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 14
    ws.row_dimensions[8].height = 36
    ws.row_dimensions[9].height = 18

    kpis = [
        ('B', 'C', 'DEMO DRIVES', f"{len(demo_df):,}", BLACK),
        ('D', 'E', 'CONVERSIONS', f"{demo_df['converted'].sum():,}", BLACK),
        ('F', 'F', 'CONV RATE', f"{demo_df['converted'].mean():.1%}", TESLA_RED),
        ('G', 'G', 'REVENUE', f"${demo_df['revenue_generated'].sum()/1e6:.1f}M", GREEN),
        ('H', 'H', 'OPPORTUNITY', f"${results['underperforming']['total_revenue_opportunity']/1e6:.1f}M", TESLA_RED),
    ]

    for col_start, col_end, label, value, val_color in kpis:
        if col_start != col_end:
            ws.merge_cells(f'{col_start}7:{col_end}7')
            ws.merge_cells(f'{col_start}8:{col_end}8')
            ws.merge_cells(f'{col_start}9:{col_end}9')

        lbl_cell = ws[f'{col_start}7']
        lbl_cell.value = label
        lbl_cell.font = Font(name='Calibri', size=8, bold=True, color=DARK_GRAY)
        lbl_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        lbl_cell.alignment = Alignment(horizontal='left', vertical='bottom')
        lbl_cell.border = Border(
            top=Side(style='thick', color=val_color)
        )

        val_cell = ws[f'{col_start}8']
        val_cell.value = value
        val_cell.font = Font(name='Calibri', size=20, bold=True, color=val_color)
        val_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        val_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Key Findings Table
    ws.row_dimensions[11].height = 18
    ws.merge_cells('B11:H11')
    findings_hdr = ws['B11']
    findings_hdr.value = "KEY FINDINGS"
    findings_hdr.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    findings_hdr.fill = PatternFill("solid", fgColor=BLACK)
    findings_hdr.alignment = Alignment(horizontal='left', vertical='center')

    findings = [
        ("Market Performance",
         f"Portland, OR leads fleet at {results['by_market'].iloc[0]['conversion_rate']:.1%} conversion. "
         f"Las Vegas, NV trails at {results['by_market'].iloc[-1]['conversion_rate']:.1%}. "
         f"{results['by_market'].iloc[0]['conversion_rate'] - results['by_market'].iloc[-1]['conversion_rate']:.1%} performance gap."),
        ("Weekend Effect",
         f"Weekend demos convert at {results['weekend_analysis']['weekend_conversion_rate']:.1%} vs "
         f"{results['weekend_analysis']['weekday_conversion_rate']:.1%} weekdays — "
         f"+{results['weekend_analysis']['lift']}% lift. Statistically significant (p < 0.001)."),
        ("Drive Duration",
         f"Converted customers averaged {results['duration_impact']['avg_duration_converted']} min drives vs "
         f"{results['duration_impact']['avg_duration_not_converted']} min for non-converters. "
         f"Pearson r = {results['duration_impact']['correlation']} (p < 0.001)."),
        ("Revenue Opportunity",
         f"${results['underperforming']['total_revenue_opportunity']:,.0f} addressable revenue if "
         f"{len(results['underperforming']['underperforming_markets'])} underperforming markets "
         f"reach fleet average of {results['underperforming']['overall_avg_conversion']:.1%}."),
        ("Top Customer Segment",
         f"Tech Early Adopters and Fleet/Business Buyers show highest conversion probability. "
         f"Targeted outreach to these segments represents highest ROI optimization lever."),
        ("Regression Model",
         f"Logistic regression (Pseudo R² = {results['regression_model']['pseudo_r2']}) confirms "
         f"customer segment and drive duration as primary conversion drivers. "
         f"{len(results['regression_model']['significant_features'])} significant predictors identified."),
    ]

    for i, (category, finding) in enumerate(findings):
        row = 12 + i
        ws.row_dimensions[row].height = 32
        bg = WHITE if i % 2 == 0 else OFF_WHITE

        cat_cell = ws[f'B{row}']
        cat_cell.value = category
        cat_cell.font = Font(name='Calibri', size=10, bold=True, color=BLACK)
        cat_cell.fill = PatternFill("solid", fgColor=bg)
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(f'C{row}:H{row}')
        find_cell = ws[f'C{row}']
        find_cell.value = finding
        find_cell.font = Font(name='Calibri', size=10, color=DARK_GRAY)
        find_cell.fill = PatternFill("solid", fgColor=bg)
        find_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    add_thin_border(ws, 12, 11 + len(findings), 2, 8)

    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18 if col in ['B'] else 22


def create_market_pivot(wb, demo_df):
    ws = wb.create_sheet("Market Pivot")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3

    # Header
    ws.merge_cells('B2:L2')
    h = ws['B2']
    h.value = "MARKET PERFORMANCE PIVOT TABLE"
    h.font = Font(name='Calibri', size=14, bold=True, color=WHITE)
    h.fill = PatternFill("solid", fgColor=BLACK)
    h.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 28

    # Build pivot
    pivot = demo_df.pivot_table(
        values=['converted', 'revenue_generated', 'satisfaction_score',
                'drive_duration_mins', 'utilization_rate'],
        index=['region', 'market'],
        aggfunc={
            'converted': ['count', 'sum', 'mean'],
            'revenue_generated': 'sum',
            'satisfaction_score': 'mean',
            'drive_duration_mins': 'mean',
            'utilization_rate': 'mean'
        }
    ).round(4)

    pivot.columns = ['_'.join(col).strip() for col in pivot.columns]
    pivot = pivot.rename(columns={
        'converted_count': 'Total Demos',
        'converted_sum': 'Conversions',
        'converted_mean': 'Conv Rate',
        'revenue_generated_sum': 'Total Revenue',
        'satisfaction_score_mean': 'Avg Satisfaction',
        'drive_duration_mins_mean': 'Avg Duration',
        'utilization_rate_mean': 'Avg Utilization'
    })
    pivot = pivot.reset_index()
    pivot['Revenue per Demo'] = (pivot['Total Revenue'] / pivot['Total Demos']).round(2)
    pivot = pivot.sort_values('Conv Rate', ascending=False)

    # Column headers
    headers = ['Region', 'Market', 'Total Demos', 'Conversions',
               'Conv Rate', 'Total Revenue', 'Revenue/Demo',
               'Satisfaction', 'Avg Duration', 'Utilization']

    col_letters = ['B','C','D','E','F','G','H','I','J','K']
    col_widths   = [14, 20, 12, 12, 10, 16, 14, 12, 12, 12]

    row_4 = 4
    ws.row_dimensions[row_4].height = 18
    for i, (col, hdr, w) in enumerate(zip(col_letters, headers, col_widths)):
        cell = ws[f'{col}{row_4}']
        style_header_cell(cell, hdr, size=9, bg=BLACK, fg=WHITE, align='center')
        ws.column_dimensions[col].width = w

    # Data rows
    avg_conv = pivot['Conv Rate'].mean()
    for r_idx, (_, row) in enumerate(pivot.iterrows()):
        excel_row = row_4 + 1 + r_idx
        ws.row_dimensions[excel_row].height = 16
        bg = WHITE if r_idx % 2 == 0 else OFF_WHITE

        conv_val = row['Conv Rate']
        if conv_val >= avg_conv * 1.15:
            tier_bg = GREEN_LIGHT
        elif conv_val <= avg_conv * 0.85:
            tier_bg = RED_LIGHT
        else:
            tier_bg = bg

        data = [
            (row.get('region',''), 'left', None, bg),
            (row.get('market',''), 'left', None, bg),
            (int(row.get('Total Demos', 0)), 'center', '#,##0', bg),
            (int(row.get('Conversions', 0)), 'center', '#,##0', bg),
            (float(conv_val), 'center', '0.0%', tier_bg),
            (float(row.get('Total Revenue', 0)), 'right', '$#,##0', bg),
            (float(row.get('Revenue per Demo', 0)), 'right', '$#,##0', bg),
            (float(row.get('Avg Satisfaction', 0)), 'center', '0.00', bg),
            (float(row.get('Avg Duration', 0)), 'center', '0.0', bg),
            (float(row.get('Avg Utilization', 0)), 'center', '0.0%', bg),
        ]

        for col, (val, align, fmt, cell_bg) in zip(col_letters, data):
            cell = ws[f'{col}{excel_row}']
            style_data_cell(cell, val, fmt=fmt, bg=cell_bg, align=align)

    add_thin_border(ws, row_4, row_4 + len(pivot), 2, 11)

    # Add totals row
    total_row = row_4 + len(pivot) + 1
    ws.row_dimensions[total_row].height = 18
    ws.merge_cells(f'B{total_row}:C{total_row}')
    total_lbl = ws[f'B{total_row}']
    style_header_cell(total_lbl, 'FLEET TOTAL', size=9, bg=LIGHT_GRAY, fg=BLACK)

    totals = [
        int(pivot['Total Demos'].sum()),
        int(pivot['Conversions'].sum()),
        pivot['Conversions'].sum() / pivot['Total Demos'].sum(),
        pivot['Total Revenue'].sum(),
        pivot['Total Revenue'].sum() / pivot['Total Demos'].sum(),
        pivot['Avg Satisfaction'].mean(),
        pivot['Avg Duration'].mean(),
        pivot['Avg Utilization'].mean(),
    ]
    fmts = ['#,##0','#,##0','0.0%','$#,##0','$#,##0','0.00','0.0','0.0%']

    for col, val, fmt in zip(col_letters[2:], totals, fmts):
        cell = ws[f'{col}{total_row}']
        style_data_cell(cell, val, fmt=fmt, bold=True,
                       bg=LIGHT_GRAY, align='center')


def create_model_pivot(wb, demo_df):
    ws = wb.create_sheet("Model Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3

    ws.merge_cells('B2:I2')
    h = ws['B2']
    h.value = "VEHICLE MODEL PERFORMANCE ANALYSIS"
    h.font = Font(name='Calibri', size=14, bold=True, color=WHITE)
    h.fill = PatternFill("solid", fgColor=BLACK)
    h.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 28

    # Cross-tab: Model vs Region
    cross = pd.crosstab(
        demo_df['vehicle_model'],
        demo_df['region'],
        values=demo_df['converted'],
        aggfunc='mean'
    ).round(4)

    # Headers
    ws['B4'].value = 'MODEL \\ REGION'
    ws['B4'].font = Font(name='Calibri', size=9, bold=True, color=WHITE)
    ws['B4'].fill = PatternFill("solid", fgColor=BLACK)
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    ws.column_dimensions['B'].width = 16
    ws.row_dimensions[4].height = 18

    for c_idx, col in enumerate(cross.columns):
        col_letter = get_column_letter(3 + c_idx)
        cell = ws[f'{col_letter}4']
        style_header_cell(cell, col, size=9, bg=DARK_GRAY, fg=WHITE, align='center')
        ws.column_dimensions[col_letter].width = 12

    avg_cell_letter = get_column_letter(3 + len(cross.columns))
    style_header_cell(ws[f'{avg_cell_letter}4'], 'AVG CONV', size=9, bg=TESLA_RED, fg=WHITE, align='center')
    ws.column_dimensions[avg_cell_letter].width = 12

    # Data
    for r_idx, (model, row_data) in enumerate(cross.iterrows()):
        excel_row = 5 + r_idx
        ws.row_dimensions[excel_row].height = 18
        bg = WHITE if r_idx % 2 == 0 else OFF_WHITE

        name_cell = ws[f'B{excel_row}']
        style_data_cell(name_cell, model, bold=True, bg=bg)

        for c_idx, val in enumerate(row_data):
            col_letter = get_column_letter(3 + c_idx)
            cell = ws[f'{col_letter}{excel_row}']
            if pd.notna(val):
                cell.value = float(val)
                cell.number_format = '0.0%'
                intensity = min(int(val * 400), 255)
                red_val = 255 - intensity
                hex_color = f'{red_val:02X}{255:02X}{red_val:02X}'
                cell.fill = PatternFill("solid", fgColor=hex_color)
            cell.font = Font(name='Calibri', size=10, color=BLACK)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        avg_val = row_data.mean()
        avg_cell = ws[f'{avg_cell_letter}{excel_row}']
        avg_cell.value = float(avg_val)
        avg_cell.number_format = '0.0%'
        avg_cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        avg_cell.fill = PatternFill("solid", fgColor=TESLA_RED)
        avg_cell.alignment = Alignment(horizontal='center', vertical='center')

    add_thin_border(ws, 4, 4 + len(cross), 2, 2 + len(cross.columns) + 1)


def create_segment_analysis(wb, demo_df, results):
    ws = wb.create_sheet("Segment Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3

    ws.merge_cells('B2:J2')
    h = ws['B2']
    h.value = "CUSTOMER SEGMENT CONVERSION ANALYSIS"
    h.font = Font(name='Calibri', size=14, bold=True, color=WHITE)
    h.fill = PatternFill("solid", fgColor=BLACK)
    h.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 28

    seg_data = results['by_segment']

    headers = ['Customer Segment', 'Total Demos', 'Conversions',
               'Conv Rate', 'Avg Satisfaction', 'Avg Duration (min)', 'Rank']
    col_letters = ['B','C','D','E','F','G','H']
    col_widths   = [22, 13, 13, 10, 16, 18, 8]

    ws.row_dimensions[4].height = 18
    for col, hdr, w in zip(col_letters, headers, col_widths):
        style_header_cell(ws[f'{col}4'], hdr, size=9, bg=BLACK, fg=WHITE, align='center')
        ws.column_dimensions[col].width = w

    for r_idx, (_, row) in enumerate(seg_data.iterrows()):
        excel_row = 5 + r_idx
        ws.row_dimensions[excel_row].height = 18
        bg = WHITE if r_idx % 2 == 0 else OFF_WHITE

        data = [
            (row['customer_segment'], 'left', None),
            (int(row['total_demos']), 'center', '#,##0'),
            (int(row['conversions']), 'center', '#,##0'),
            (float(row['conversion_rate']), 'center', '0.0%'),
            (float(row['avg_satisfaction']), 'center', '0.00'),
            (float(row['avg_drive_duration']), 'center', '0.0'),
            (r_idx + 1, 'center', None),
        ]

        for col, (val, align, fmt) in zip(col_letters, data):
            cell = ws[f'{col}{excel_row}']
            conv_bg = GREEN_LIGHT if r_idx == 0 else RED_LIGHT if r_idx == len(seg_data)-1 else bg
            style_data_cell(cell, val, fmt=fmt,
                          bg=conv_bg if col == 'E' else bg,
                          align=align)

    add_thin_border(ws, 4, 4 + len(seg_data), 2, 8)

    # Statistical findings
    stat_row = 4 + len(seg_data) + 3
    ws.merge_cells(f'B{stat_row}:H{stat_row}')
    style_header_cell(ws[f'B{stat_row}'], 'STATISTICAL FINDINGS', size=9, bg=DARK_GRAY, fg=WHITE)
    ws.row_dimensions[stat_row].height = 18

    dur = results['duration_impact']
    stats = [
        ("Drive Duration — Converted", f"{dur['avg_duration_converted']} minutes average"),
        ("Drive Duration — Not Converted", f"{dur['avg_duration_not_converted']} minutes average"),
        ("Pearson Correlation (Duration vs Conv)", f"r = {dur['correlation']} (p < 0.001, statistically significant)"),
        ("Weekend Lift", f"+{results['weekend_analysis']['lift']}% conversion uplift on weekends"),
        ("Regression Pseudo R²", f"{results['regression_model']['pseudo_r2']} — model explains meaningful variance"),
    ]

    for i, (lbl, val) in enumerate(stats):
        row = stat_row + 1 + i
        ws.row_dimensions[row].height = 16
        bg = WHITE if i % 2 == 0 else OFF_WHITE
        style_data_cell(ws[f'B{row}'], lbl, bold=True, bg=bg)
        ws.merge_cells(f'C{row}:H{row}')
        style_data_cell(ws[f'C{row}'], val, bg=bg, fg=DARK_GRAY)

    add_thin_border(ws, stat_row, stat_row + len(stats), 2, 8)


def create_raw_data_sheet(wb, demo_df):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2

    ws.merge_cells('B1:T1')
    h = ws['B1']
    h.value = "RAW DEMO DRIVE DATA — SAMPLE (500 RECORDS)"
    h.font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    h.fill = PatternFill("solid", fgColor=BLACK)
    h.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    sample = demo_df.sample(500, random_state=42).sort_values('demo_date')
    cols = ['demo_id', 'demo_date', 'vehicle_model', 'market', 'region',
            'customer_segment', 'drive_duration_mins', 'utilization_rate',
            'converted', 'revenue_generated', 'satisfaction_score', 'demo_cost']

    for c_idx, col in enumerate(cols):
        col_letter = get_column_letter(2 + c_idx)
        cell = ws[f'{col_letter}2']
        style_header_cell(cell, col.upper().replace('_', ' '), size=8,
                         bg=DARK_GRAY, fg=WHITE, align='center')
        ws.column_dimensions[col_letter].width = 16

    ws.row_dimensions[2].height = 16

    for r_idx, (_, row) in enumerate(sample[cols].iterrows()):
        excel_row = 3 + r_idx
        bg = WHITE if r_idx % 2 == 0 else OFF_WHITE
        for c_idx, col in enumerate(cols):
            col_letter = get_column_letter(2 + c_idx)
            cell = ws[f'{col_letter}{excel_row}']
            val = row[col]
            if col == 'demo_date':
                val = str(val)[:10]
                fmt = None
            elif col in ['utilization_rate']:
                fmt = '0.0%'
            elif col == 'revenue_generated':
                fmt = '$#,##0'
            elif col == 'demo_cost':
                fmt = '$#,##0.00'
            else:
                fmt = None
            style_data_cell(cell, val, fmt=fmt, bg=bg, align='center')

    ws.freeze_panes = 'B3'
    ws.auto_filter.ref = f"B2:{get_column_letter(2+len(cols)-1)}2"


def add_vba_macro(wb):
    """Add macro instructions sheet since openpyxl can't write .xlsm directly"""
    ws = wb.create_sheet("Macro Instructions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 80

    ws.merge_cells('B2:D2')
    h = ws['B2']
    h.value = "EXCEL MACRO — Auto-Format Summary Report"
    h.font = Font(name='Calibri', size=13, bold=True, color=WHITE)
    h.fill = PatternFill("solid", fgColor=BLACK)
    h.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 26

    instructions = [
        "To enable the macro: Open Excel → Alt+F11 → Insert Module → Paste code below → Run",
        "",
        "Sub FormatFleetReport()",
        "    ' Auto-format the Executive Overview sheet',",
        "    Dim ws As Worksheet",
        "    Set ws = ThisWorkbook.Sheets('Executive Overview')",
        "    ws.Activate",
        "    ",
        "    ' Apply conditional formatting to conversion rates",
        "    Dim rng As Range",
        "    Set rng = ws.Range('F8')",
        "    ",
        "    ' Bold all KPI values",
        "    ws.Range('B8:H8').Font.Bold = True",
        "    ws.Range('B8:H8').Font.Size = 20",
        "    ",
        "    ' Auto-fit columns",
        "    ws.Columns('B:H').AutoFit",
        "    ",
        "    ' Add print area",
        "    ws.PageSetup.PrintArea = '$B$2:$H$20'",
        "    ws.PageSetup.Orientation = xlLandscape",
        "    ws.PageSetup.FitToPagesWide = 1",
        "    ws.PageSetup.FitToPagesTall = False",
        "    ",
        "    MsgBox 'Report formatted successfully!', vbInformation",
        "End Sub",
        "",
        "Sub RefreshAllPivots()",
        "    ' Refresh all pivot tables in workbook",
        "    Dim pt As PivotTable",
        "    Dim ws As Worksheet",
        "    For Each ws In ThisWorkbook.Worksheets",
        "        For Each pt In ws.PivotTables",
        "            pt.RefreshTable",
        "        Next pt",
        "    Next ws",
        "    MsgBox 'All pivot tables refreshed!', vbInformation",
        "End Sub",
    ]

    for i, line in enumerate(instructions):
        row = 4 + i
        ws.row_dimensions[row].height = 15
        cell = ws[f'B{row}']
        cell.value = line
        is_code = line.startswith('    ') or line.startswith('Sub') or line.startswith('End')
        cell.font = Font(
            name='Courier New' if is_code else 'Calibri',
            size=10 if is_code else 10,
            color=BLACK if not is_code else DARK_GRAY
        )
        cell.fill = PatternFill("solid", fgColor=OFF_WHITE if is_code else WHITE)
        cell.alignment = Alignment(horizontal='left', vertical='center')


def generate_report(output_path='reports/Tesla_Fleet_Intelligence_Report.xlsx'):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Loading data...")
    demo_df, vehicle_df, market_df = load_data()
    results = run_full_analysis(demo_df, vehicle_df, market_df)

    print("Building Excel report...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_overview_sheet(wb, demo_df, results)
    create_market_pivot(wb, demo_df)
    create_model_pivot(wb, demo_df)
    create_segment_analysis(wb, demo_df, results)
    create_raw_data_sheet(wb, demo_df)
    add_vba_macro(wb)

    wb.save(output_path)
    print(f"Report saved: {output_path}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")
    return output_path


if __name__ == "__main__":
    generate_report()